from Input.preprocess import generate_all_stations, reset_stations, generate_pattern_stations, reset_cap_stations
from vehicle import Vehicle
from Output.save_to_excel import *
from Simulation.BSS_environment import Environment
import copy
import pandas as pd
from openpyxl import load_workbook


start_hour = 7
no_stations = 200
branching = 3
subproblem_scenarios = 10
simulation_time = 960  # 7 am to 11 pm
stations = generate_all_stations(start_hour, no_stations)
stations[4].depot = True


def get_criticality_weights():
    # w_drive, w_dev, w_viol, w_net
    vals = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
    weights = list()
    for val1 in vals:
        w_drive = val1
        for val2 in vals:
            if w_drive + val2 <= 1:
                w_dev = val2
            else:
                break
            for val3 in vals:
                if w_drive + w_dev + val3 <= 1:
                    w_viol = val3
                else:
                    break
                w_flat = 1 - w_drive - w_dev - w_viol
                weights.append((w_drive, w_dev, w_viol, w_flat))
    return weights


def get_weight_combination_reduced():
    # W_V, W_R, W_D, W_VN, W_VL
    weights = list()
    vals_d = [0.1, 0.2, 0.3, 0.4]
    vals_v = [0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
    vals_n = [0.5, 0.6, 0.7, 0.8, 0.9, 1]
    for val1 in vals_v:
        W_V = val1
        for val2 in vals_d:
            if W_V + val2 <= 1:
                W_D = val2
            else:
                break
            W_R = 1 - W_D - W_V
            for val3 in vals_n:
                W_N = val3
                W_L = 1 - W_N
                weights.append((W_V, W_R, W_D, W_N, W_L))
    return weights


def weight_analysis(choice):
    all_sets = get_criticality_weights()
    env = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios)

    # Generating 10 scenarios
    scenarios = [env.generate_trips(simulation_time//60, gen=True) for i in range(1)]

    # Create excel writer
    writer = pd.ExcelWriter("Output/output_weights_" + choice + ".xlsx", engine='openpyxl')
    book = load_workbook("Output/output_weights_" + choice + ".xlsx")
    writer.book = book

    base_viol = list()
    for j in range(len(scenarios)):
        reset_stations(stations)
        init_base_stack = [copy.copy(trip) for trip in scenarios[j]]
        sim_base = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios,
                               trigger_start_stack=init_base_stack, memory_mode=True)
        sim_base.run_simulation()
        base_viol.append(sim_base)

    for i in range(len(all_sets)):
        for j in range(len(scenarios)):
            reset_stations(stations)
            init_stack = [copy.copy(trip) for trip in scenarios[j]]
            v = Vehicle(init_battery_load=40, init_charged_bikes=20, init_flat_bikes=0, current_station=stations[0],
                          id=0)
            sim_env = Environment(start_hour, simulation_time, stations, [v], branching, subproblem_scenarios,
                                  trigger_start_stack=init_stack, memory_mode=True, crit_weights=all_sets[i])
            sim_env.run_simulation()
            save_weight_output(i+1, j+1, sim_env, base_viol[j].total_starvations, base_viol[j].total_congestions, writer)


def strategy_analysis(scen, veh, run):
    # Create excel writer
    writer = pd.ExcelWriter("Output/runtime_" + run + ".xlsx", engine='openpyxl')
    book = load_workbook("Output/runtime_" + run + ".xlsx")
    writer.book = book

    vehicles = list()
    for i in range(veh):
        vehicles.append(Vehicle(init_battery_load=40, init_charged_bikes=0, init_flat_bikes=0,
                                current_station=stations[i], id=i))
    env = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios)

    # Generating scenarios
    scenarios = [env.generate_trips(simulation_time//60, gen=True) for i in range(scen)]

    scenario = 1
    for sc in scenarios:
        reset_stations(stations)
        # Base
        init_base_stack = [copy.copy(trip) for trip in sc]
        sim_base = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios,
                               trigger_start_stack=init_base_stack, memory_mode=True)
        sim_base.run_simulation()
        reset_stations(stations)

        """
        # Greedy
        init_greedy_stack = [copy.copy(trip) for trip in sc]
        vehicles_greedy = [copy.copy(veh) for veh in vehicles]
        sim_greedy = Environment(start_hour, simulation_time, stations, vehicles_greedy, branching,
                                 subproblem_scenarios, trigger_start_stack=init_greedy_stack, memory_mode=True,
                                 greedy=True, writer=writer1)
        sim_greedy.run_simulation()
        """

        # Crit off
        reset_stations(stations)
        init_crit_stack = [copy.copy(trip) for trip in sc]
        vehicles_crit = [copy.copy(veh) for veh in vehicles]
        crit_env = Environment(start_hour, simulation_time, stations, vehicles_crit, branching,
                               subproblem_scenarios, trigger_start_stack=init_crit_stack, memory_mode=True,
                               criticality=False)

        # 10 subproblem
        reset_stations(stations)
        init_heur_stack = [copy.copy(trip) for trip in sc]
        vehicles_heur = [copy.copy(veh) for veh in vehicles]
        sim_heur = Environment(start_hour, simulation_time, stations, vehicles_heur, branching,
                                 subproblem_scenarios, trigger_start_stack=init_heur_stack, memory_mode=True,
                               criticality=True)
        sim_heur.run_simulation()
        save_vehicle_output(scenario, veh, sim_heur, sim_base, sim_base, writer, crit_env, alfa=1)
        scenario += 1


def first_step():
    # Create excel writer
    writer = pd.ExcelWriter("Output/output.xlsx", engine='openpyxl')
    book = load_workbook("Output/output.xlsx")
    writer.book = book
    v = Vehicle(init_battery_load=40, init_charged_bikes=10, init_flat_bikes=0
                , current_station=stations[0], id=0)
    sim_env = Environment(start_hour, simulation_time, stations, [v], branching, subproblem_scenarios, writer=writer)
    sim_env.run_simulation()


def runtime_analysis(run):
    # Create excel writer
    writer = pd.ExcelWriter("Output/runtime_"+run+".xlsx", engine='openpyxl')
    book = load_workbook("Output/runtime_"+run+".xlsx")
    writer.book = book
    stations = generate_pattern_stations(200)
    for sub_sc in [1, 10, 20, 30]:
        for no_vehicles in [1, 2, 3, 4, 5]:
                vehicles = list()
                for i in range(no_vehicles):
                    vehicles.append(Vehicle(init_battery_load=40, init_charged_bikes=8, init_flat_bikes=8, bike_cap=25,
                                            current_station=stations[i], id=i))
                for no_stations in [20, 50, 100, 150, 200]:
                    sta = stations[:no_stations]
                    sim_env = Environment(start_hour, simulation_time, sta, vehicles, branching, sub_sc,
                                          memory_mode=True)
                    sim_env.run_simulation()
                    time = sim_env.event_times[0]
                    save_time_output(no_stations, branching, sub_sc, no_vehicles, time, writer)
                    reset_stations(stations)


def vehicle_analysis(days, veh, run):
    env = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios)

    # Generating days
    days = [env.generate_trips(simulation_time // 60, gen=True) for i in range(days)]

    # Create excel writer
    writer = pd.ExcelWriter("Output/runtime_" + run + ".xlsx", engine='openpyxl')
    book = load_workbook("Output/runtime_" + run + ".xlsx")
    writer.book = book

    base_envs = list()
    for j in range(len(days)):
        reset_stations(stations)
        init_base_stack = [copy.copy(trip) for trip in days[j]]
        sim_base = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios,
                               trigger_start_stack=init_base_stack, memory_mode=True)
        sim_base.run_simulation()
        base_envs.append(sim_base)

    for d in range(len(days)):
        for n_veh in range(1, veh+1):
            vehicles = list()
            for k in range(n_veh):
                vehicles.append(Vehicle(init_battery_load=40, init_charged_bikes=0, init_flat_bikes=0,
                                        current_station=stations[k], id=k))
            reset_stations(stations)
            init_heur_stack = [copy.copy(trip) for trip in days[d]]
            vehicles_heur = [copy.copy(veh) for veh in vehicles]
            sim_heur = Environment(start_hour, simulation_time, stations, vehicles_heur, branching,
                                   subproblem_scenarios, trigger_start_stack=init_heur_stack, memory_mode=True,
                                   criticality=True)
            sim_heur.run_simulation()
            save_vary_vehicle_output(d+1, n_veh, sim_heur, base_envs[d], writer)


def fleet_analysis(days, run):
    env = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios)

    # Generating days
    days = [env.generate_trips(simulation_time // 60, gen=True) for i in range(days)]

    # Create excel writer
    writer = pd.ExcelWriter("Output/runtime_" + run + ".xlsx", engine='openpyxl')
    book = load_workbook("Output/runtime_" + run + ".xlsx")
    writer.book = book

    base_envs = list()
    for j in range(len(days)):
        reset_stations(stations)
        init_base_stack = [copy.copy(trip) for trip in days[j]]
        sim_base = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios,
                               trigger_start_stack=init_base_stack, memory_mode=True)
        sim_base.run_simulation()
        base_envs.append(sim_base)

    for d in range(len(days)):
        for n_bat in range(0, 6):
            vehicles = list()
            for k in range(n_bat):
                vehicles.append(Vehicle(init_battery_load=40, init_charged_bikes=0, init_flat_bikes=0,
                                        current_station=stations[k], id=k, bike_cap=0))
            for n_rb in range(0, 6):
                veh1 = list()
                veh1 += vehicles
                for l in range(n_rb):
                    index = l + n_bat
                    if l+n_bat == 4:
                        index = 15
                    veh1.append(Vehicle(init_battery_load=0, init_charged_bikes=0, init_flat_bikes=0,
                                        current_station=stations[index], id=l+n_bat, bat_cap=0))
                reset_stations(stations)
                init_heur_stack = [copy.copy(trip) for trip in days[d]]
                vehicles_heur = [copy.copy(veh) for veh in veh1]
                sim_heur = Environment(start_hour, simulation_time, stations, vehicles_heur, branching,
                                       subproblem_scenarios, trigger_start_stack=init_heur_stack, memory_mode=True,
                                       criticality=True)
                sim_heur.run_simulation()
                save_fleet_output(d+1, n_rb, n_bat, sim_heur, base_envs[d], writer)


def station_cap(no_days, run):
    env = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios)

    # Generating days
    days = [env.generate_trips(simulation_time // 60, gen=True) for i in range(no_days)]

    # Create excel writer
    writer = pd.ExcelWriter("Output/runtime_" + run + ".xlsx", engine='openpyxl')
    book = load_workbook("Output/runtime_" + run + ".xlsx")
    writer.book = book

    vehicles = list()
    for i in range(5):
        vehicles.append(Vehicle(init_battery_load=40, init_charged_bikes=0, init_flat_bikes=0,
                                current_station=stations[i], id=i))

    for d in range(no_days):
        for m in [1, 1.25, 1.5, 1.75, 2, 2.25, 2.5, 2.75, 3]:
            reset_cap_stations(stations, m)
            init_base_stack = [copy.copy(trip) for trip in days[d]]
            sim_base = Environment(start_hour, simulation_time, stations, list(), branching, subproblem_scenarios,
                                   trigger_start_stack=init_base_stack, memory_mode=True)
            sim_base.run_simulation()

            reset_cap_stations(stations, m)
            init_heur_stack = [copy.copy(trip) for trip in days[d]]
            vehicles_heur = [copy.copy(veh) for veh in vehicles]
            sim_heur = Environment(start_hour, simulation_time, stations, vehicles_heur, branching,
                                   subproblem_scenarios, trigger_start_stack=init_heur_stack, memory_mode=True,
                                   criticality=True)
            sim_heur.run_simulation()

            save_station_cap_output(d+1, sim_heur, sim_base, m, writer)


if __name__ == '__main__':
    print("w: weight analysis, c: strategy comparison, r: runtime analysis, fs: first step analysis, v: vehicles")
    choice = input('Choose action: ')
    if choice == 'w1':
        weight_analysis(choice)
    elif choice == 'v':
        scenarios = input('Number of days:')
        vehicles = input('Number of vehicles:')
        run = input('run number:')
        vehicle_analysis(int(scenarios), int(vehicles), run)
    elif choice == 'sc':
        days = input('Number of days:')
        run = input('run number:')
        station_cap(int(days), run)
    elif choice == 'vf':
        days = input('Number of days:')
        run = input('run number:')
        fleet_analysis(int(days), run)
    elif choice == 'c':
        days = input('Number of days:')
        vehicles = input('Number of vehicles:')
        run = input('Run number:')
        strategy_analysis(int(days), int(vehicles), run)
    elif choice == 'fs':
        first_step()
    elif choice == 'r':
        run = input('run number:')
        runtime_analysis(run)
    else:
        print("No analysis")
