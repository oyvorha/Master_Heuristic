import pandas as pd
from openpyxl import load_workbook

df_keys = []


def save_time_output(no_stations, branching, scenarios, no_vehicles, time):
    book = load_workbook("Output/output.xlsx")
    writer = pd.ExcelWriter("Output/output.xlsx", engine='openpyxl')
    writer.book = book

    df = pd.DataFrame(columns=['Stations', 'Vehicles', 'Branching', 'Scenarios', 'Time'])

    new_row = {'Stations': no_stations, 'Vehicles': no_vehicles, 'Branching': branching, 'Scenarios': scenarios,
               'Time': time}

    time_df = df.append(new_row, ignore_index=True)

    if 'solution_time' in df_keys:
        start_row = writer.sheets['solution_time'].max_row
        time_df.to_excel(writer, startrow=start_row, index=False, header=False, sheet_name='solution_time')
        writer.save()
    else:
        time_df.to_excel(writer, index=False, sheet_name='solution_time')
        writer.save()
        df_keys.append('solution_time')


def save_weight_output(set_id, scenario, env, base_s, base_c, choice):
    book_w = load_workbook("Output/output_weights_" + choice +".xlsx")
    writer_w = pd.ExcelWriter("Output/output_weights_" + choice +".xlsx", engine='openpyxl')
    writer_w.book = book_w

    df = pd.DataFrame(columns=['Weight set', 'W_V', 'W_R', 'W_D', 'W_VN', 'W_VL', 'Scenario', 'Total_requests',
                               'Base starvations', 'Base congestions', 'Starvations', 'Congestions'])

    new_row = {'Weight set': set_id, 'W_V': env.weights[0], 'W_R': env.weights[1], 'W_D': env.weights[2],
               'W_VN': env.weights[3], 'W_VL': env.weights[4], 'Scenario': scenario,
               'Total_requests': env.total_gen_trips, 'Base starvations': base_s, 'Base congestions': base_c,
               'Starvations': env.total_starvations, 'Congestions': env.total_congestions}

    weight_df = df.append(new_row, ignore_index=True)

    if 'Weight_simulation' in df_keys:
        start_row = writer_w.sheets['Weight_simulation'].max_row
        weight_df.to_excel(writer_w, startrow=start_row, index=False, header=False, sheet_name='Weight_simulation')
        writer_w.save()
    else:
        weight_df.to_excel(writer_w, index=False, sheet_name='Weight_simulation')
        writer_w.save()
        df_keys.append('Weight_simulation')


def save_comparison_output(scenario, sim_heur, base_s, base_c, greedy_s, greedy_c):
    book_c = load_workbook("Output/output_compare.xlsx")
    writer_c = pd.ExcelWriter("Output/output_compare.xlsx", engine='openpyxl')
    writer_c.book = book_c

    df = pd.DataFrame(columns=['Scenario', 'Total_requests', 'Base starvations', 'Base congestions',
                               'Greedy starvations', 'Greedy congestions', 'Heuristic starvations',
                               'Heuristic congestions'])

    new_row = {'Scenario': scenario, 'Total_requests': sim_heur.total_gen_trips, 'Base starvations': base_s,
               'Base congestions': base_c, 'Greedy starvations': greedy_s, 'Greedy congestions': greedy_c,
               'Heuristic starvations': sim_heur.total_starvations, 'Heuristic congestions': sim_heur.total_congestions}

    weight_df = df.append(new_row, ignore_index=True)

    if 'Compare_strategies' in df_keys:
        start_row = writer_c.sheets['Compare_strategies'].max_row
        weight_df.to_excel(writer_c, startrow=start_row, index=False, header=False, sheet_name='Compare_strategies')
        writer_c.save()
    else:
        weight_df.to_excel(writer_c, index=False, sheet_name='Compare_strategies')
        writer_c.save()
        df_keys.append('Compare_strategies')
